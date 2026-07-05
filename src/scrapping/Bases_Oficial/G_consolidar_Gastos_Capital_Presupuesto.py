import pandas as pd
import os
import locale
from openpyxl.styles import Font, PatternFill, Alignment

def obtener_separador_formula():

    try:
        configuracion = locale.localeconv()

        if configuracion['decimal_point'] == ',':
            return ';'

        return ','

    except:
        return ';'


def aplicar_formato_numerico(worksheet):

    rangos = [
        ('F', 'R'),
        ('X', 'AL')
    ]

    ultima_fila = worksheet.max_row

    for col_inicio, col_fin in rangos:

        for fila in range(2, ultima_fila + 1):

            for celda in worksheet[f'{col_inicio}{fila}:{col_fin}{fila}'][0]:

                if celda.value is not None:
                    celda.number_format = '#,##0.00'


def aplicar_estilo_encabezados(worksheet):

    relleno = PatternFill(
        fill_type='solid',
        start_color='C0C0C0',
        end_color='C0C0C0'
    )

    fuente = Font(bold=True)

    alineacion = Alignment(
        horizontal='center',
        vertical='center'
    )

    for celda in worksheet[1]:

        if celda.value not in [None, ""]:

            celda.fill = relleno
            celda.font = fuente
            celda.alignment = alineacion


def obtener_ultima_fila(ws, columna='A'):

    for fila in range(ws.max_row, 1, -1):

        valor = ws[f'{columna}{fila}'].value

        if valor not in (None, ''):
            return fila

    return 1


def agregar_validaciones(worksheet):

    separador = obtener_separador_formula()

    worksheet["AN1"] = "VALIDACION SIGLAS"
    worksheet["AO1"] = "VALIDACION RUBROS"

    worksheet["AP1"] = f'=COUNTIF(AN:AN{separador}FALSE)'
    worksheet["AQ1"] = f'=COUNTIF(AO:AO{separador}FALSE)'

    ultima_fila = obtener_ultima_fila(worksheet, 'A')

    for fila in range(2, ultima_fila + 1):

        worksheet[f"AN{fila}"] = f"=A{fila}=T{fila}"
        worksheet[f"AO{fila}"] = f"=C{fila}=U{fila}"


# -----------------------------
# PROCESO PRINCIPAL
# -----------------------------
def consolidar_gk_presupuesto(ruta_principal):

    ruta_ejecucion = rf"{ruta_principal}\EJECUCION\Gastos_Capital_Ejecucion_Presupuesto.xlsx"
    ruta_marco = rf"{ruta_principal}\MARCO\Gastos_Capital_Formulacion_Presupuesto.xlsx"
    ruta_destino = rf"{ruta_principal}\VALIDACION GASTO CAPITAL\Validación_Gastos_Capital_Presupuesto.xlsx"

    try:

        print("Leyendo archivos fuente...")

        df_marco = pd.read_excel(ruta_marco, sheet_name=1)

        df_ejecucion = pd.read_excel(ruta_ejecucion, sheet_name=0)

        df_marco = df_marco.dropna(how='all').dropna(axis=1, how='all')
        df_ejecucion = df_ejecucion.dropna(how='all').dropna(axis=1, how='all')

        reglas = {
            "PROGRAMA DE INVERSIONES": ("1", "PGI"),
            "PROYECTOS DE INVERSION": ("2", "PI"),
            "GASTOS DE CAPITAL NO LIGADOS A PROYECTOS": ("3", "NL"),
            "INVERSION FINANCIERA": ("4", "IF"),
            "OTROS": ("5", "OT"),
            "TOTAL GASTOS DE CAPITAL": ("6", "GC")
        }

        codigos = []

        codigo_actual = ""
        ultima_regla_num = 0
        ultima_regla_texto = None

        for valor in df_marco["NOMBRE_GASTO"]:

            texto = str(valor).strip().upper()

            if texto in reglas:

                numero, abreviatura = reglas[texto]

                numero_int = int(numero)

                if ultima_regla_num == 6:
                    siguiente_esperado = 1
                else:
                    siguiente_esperado = ultima_regla_num + 1

                # Si la regla está en el orden correcto
                if numero_int == siguiente_esperado:

                    # Si se repite consecutivamente
                    if texto == ultima_regla_texto:

                        codigos.append(abreviatura)

                    else:

                        codigos.append(numero)

                    codigo_actual = abreviatura

                    # Solo avanzar flujo cuando NO es repetido
                    if texto != ultima_regla_texto:

                        ultima_regla_num = numero_int

                else:
                    # Regla fuera de secuencia
                    codigos.append(codigo_actual)

                ultima_regla_texto = texto

            else:
                codigos.append(codigo_actual)

                ultima_regla_texto = None

        posicion = df_marco.columns.get_loc("SIGLAS") + 1

        df_marco.insert(posicion, "CODIGO", codigos)

        columnas = [
            'SIGLAS',
            'CODIGO',
            'NOMBRE_GASTO',
            'TIPO_PROYECTO',
            'CODIGO_UNICO',
            'COSTO_INVERSION'
        ]

        columnasEjec = [
            'SIGLAS',
            'NOMBRE_GASTO',
            'TIPO_PROYECTO',
            'CODIGO_UNICO',
            'COSTO_INVERSION',
            'MONTO_INICIAL_AL_MES_CORTE',
            'EJEC_AÑO_ANTERIOR'
        ]

        columnas_montos = [f'MONTO{i}' for i in range(1, 13)]

        columnas_finales = columnas + columnas_montos
        columnas_finalesEjec = columnasEjec + columnas_montos

        df_marco = df_marco[columnas_finales]

        df_ejecucion = df_ejecucion[
            [c for c in columnas_finalesEjec if c != "CODIGO"]
        ]

        espacio = pd.DataFrame({
            ' ': [''] * max(len(df_marco), len(df_ejecucion))
        })

        df_final = pd.concat(
            [
                df_marco.reset_index(drop=True),
                espacio,
                df_ejecucion.reset_index(drop=True)
            ],
            axis=1
        )

        dir_destino = os.path.dirname(ruta_destino)
        os.makedirs(dir_destino, exist_ok=True)
        
        import tempfile
        fd, temp_name = tempfile.mkstemp(prefix="Validacion_GC_PRE_", suffix=".xlsx", dir=dir_destino)
        os.close(fd)

        with pd.ExcelWriter(temp_name, engine='openpyxl') as writer:

            df_final.to_excel(
                writer,
                sheet_name='Validacion_Comparativa',
                index=False
            )

            worksheet = writer.sheets['Validacion_Comparativa']

            agregar_validaciones(worksheet)

            aplicar_formato_numerico(worksheet)

            aplicar_estilo_encabezados(worksheet)

        os.replace(temp_name, ruta_destino)
        print(f"Éxito: Archivo generado en:\n{ruta_destino}")

    except KeyError as e:
        print(f"ERROR: No existe la columna {e}")

    except PermissionError:
        print("ERROR: El archivo de salida está abierto.")

    except Exception as e:
        print(f"Error inesperado: {e}")