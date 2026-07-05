import os
from openpyxl import load_workbook

def obtener_ultima_fila(ws, columna='A', fila_minima=5):
    for fila in range(ws.max_row, fila_minima - 1, -1):
        valor = ws[f"{columna}{fila}"].value
        if valor is not None and str(valor).strip() != '':
            return fila
    return fila_minima


def formato_deposiciones_colocaciones(ruta_principal):
    ruta_destino = os.path.join(ruta_principal, "Base FONAFE WEB al mes.xlsm")

    if not os.path.exists(ruta_destino):
        print("No existe archivo destino.")
        return

    wb = None
    import tempfile
    
    try:
        print("Abriendo archivo destino en segundo plano (openpyxl)...")
        wb = load_workbook(ruta_destino, keep_vba=True)
        ws = wb["Depósitos y Colocaciones"]

        ultima_fila = obtener_ultima_fila(ws, columna='A', fila_minima=5)
        print(f"Última fila detectada para dar formato: {ultima_fila}")

        print("Aplicando formatos numéricos y de fecha...")
        
        # Formatos cacheados en variables para optimizar velocidad del bucle
        formato_numero = '#,##0.00'
        formato_fecha = 'dd/mm/yyyy'
        columnas_num = ['H', 'I', 'J']

        for fila in range(5, ultima_fila + 1):
            # Formato monetario / numérico
            for col_letra in columnas_num:
                ws[f"{col_letra}{fila}"].number_format = formato_numero
            
            # Formato de fecha
            ws[f"L{fila}"].number_format = formato_fecha

        print("Guardando cambios en el archivo de forma segura...")
        print(">>> ATENCIÓN: El guardado puede demorar unos minutos. Por favor, espera... <<<")
        
        fd, temp_file = tempfile.mkstemp(prefix="Base_FONAFE_", suffix=".xlsm", dir=os.path.dirname(ruta_destino))
        os.close(fd)
        
        wb.save(temp_file)
        wb.close()
        wb = None
        os.replace(temp_file, ruta_destino)
        
        print(f"\n✓ ¡Formatos aplicados correctamente hasta la fila {ultima_fila}!")

    except PermissionError:
        print("\n✗ ERROR: El archivo Excel destino está abierto. Ciérrelo antes de ejecutar.")
    except Exception as e:
        print(f"\n✗ ERROR Inesperado en el hilo de formateo: {e}")
    finally:
        print("Liberando recursos de la memoria...")
        if wb:
            try:
                wb.close()
            except:
                pass

    print("Proceso de diseño y formatos openpyxl finalizado.")
