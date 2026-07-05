import os
import time
import pandas as pd
from openpyxl import load_workbook

import os
import time
import pandas as pd
from openpyxl import load_workbook
import tempfile

def actualizar_gastos_capital(ruta_principal: str, max_workers: int = 8):
    # Nota: max_workers se mantiene en la firma por compatibilidad si algo lo pasaba, 
    # pero el procesamiento será estrictamente secuencial.
    ruta = os.path.join(ruta_principal, "MARCO")
    CATEGORIA_ORDEN = {"M": 1, "I": 2, "P": 3, "S": 4, "3": 5}

    def procesar_archivo(ruta_archivo: str) -> dict:
        archivo = os.path.basename(ruta_archivo)
        resultado = {
            "archivo": archivo,
            "ok": False,
            "hoja": "",
            "eliminadas": 0,
            "error": ""
        }

        for intento in range(3):
            try:
                df = pd.read_excel(ruta_archivo, sheet_name=0, header=0, engine="openpyxl").copy()

                col_empresa      = df.columns[2]   # columna C
                col_modificacion = df.columns[5]   # columna F

                df[col_empresa] = df[col_empresa].astype(str).str.strip()
                df[col_modificacion] = df[col_modificacion].astype(str).str.strip()

                df["_orden"] = (
                    df[col_modificacion]
                    .map(CATEGORIA_ORDEN)
                    .fillna(0)
                )

                idx_max = df.groupby(col_empresa)["_orden"].idxmax()

                mejor_mod = (
                    df.loc[idx_max, [col_empresa, col_modificacion]]
                    .set_index(col_empresa)[col_modificacion]
                    .copy()
                )

                mascara = df.apply(
                    lambda r: mejor_mod.get(r[col_empresa], "") == r[col_modificacion],
                    axis=1
                )

                eliminadas = int((~mascara).sum())

                df_filtrado = (
                    df[mascara]
                    .drop(columns=["_orden"])
                    .reset_index(drop=True)
                    .copy()
                )

                wb = load_workbook(ruta_archivo)
                nombre_copia = f"Gastos_Capital_{len(wb.worksheets) + 1}"
                ws_copia = wb.create_sheet(title=nombre_copia)

                # Encabezados
                ws_copia.append(list(df_filtrado.columns))

                # Datos
                for row in df_filtrado.itertuples(index=False, name=None):
                    ws_copia.append(list(row))

                dir_ruta = os.path.dirname(ruta_archivo)
                fd, temp_file = tempfile.mkstemp(prefix="Gasto_Capital_", suffix=".xlsx", dir=dir_ruta)
                os.close(fd)
                
                wb.save(temp_file)
                wb.close()
                os.replace(temp_file, ruta_archivo)

                resultado.update({
                    "ok": True,
                    "hoja": nombre_copia,
                    "eliminadas": eliminadas,
                    "error": ""
                })
                break 

            except PermissionError:
                time.sleep(0.5)
                resultado["error"] = "Archivo bloqueado temporalmente por I/O"
            except Exception as e:
                resultado["error"] = str(e)
                break

        return resultado

    if not os.path.isdir(ruta):
        print(f"La carpeta no existe: {ruta}")
        return

    archivos_unicos = list(set([
        os.path.join(ruta, f)
        for f in os.listdir(ruta)
        if f.lower().endswith((".xlsx", ".xls", ".xlsm"))
        and f.startswith("Gastos_Capital")
        and not f.startswith("~$")
    ]))

    if not archivos_unicos:
        print("No se encontraron archivos Gastos_Capital válidos.")
        return

    total = len(archivos_unicos)
    print(f"Archivos Gastos_Capital únicos: {total} (Ejecución estrictamente secuencial)\n")

    t_inicio = time.perf_counter()
    completados = 0
    errores = 0

    # Procesamiento secuencial
    for ruta_archivo in archivos_unicos:
        res = procesar_archivo(ruta_archivo)
        completados += 1
        estado = f"[{completados}/{total}]"

        if res["ok"]:
            print(
                f"  {estado} OK  {res['archivo']}  "
                f"-> '{res['hoja']}' | eliminadas: {res['eliminadas']}"
            )
        else:
            errores += 1
            print(
                f"  {estado} ERR {res['archivo']}  "
                f"->  {res['error']}"
            )

    t_fin = time.perf_counter()
    print(
        f"\nCompletado en {t_fin - t_inicio:.2f}s | "
        f"OK: {completados - errores} | "
        f"Errores: {errores}"
    )

