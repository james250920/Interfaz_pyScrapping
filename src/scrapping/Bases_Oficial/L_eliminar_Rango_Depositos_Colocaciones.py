import os
from openpyxl import load_workbook

def limpiar_hojas_excel(ruta_principal):
    ruta_archivo = os.path.join(ruta_principal, "Base FONAFE WEB al mes.xlsm")
    
    try:
        import tempfile
        print("Abriendo archivo Excel...")
        wb = load_workbook(ruta_archivo, keep_vba=True)

        hojas_a_limpiar = ["Depósitos y Colocaciones"]
        rango_borrar = "A5:L6000"

        for nombre_hoja in hojas_a_limpiar:
            try:
                if nombre_hoja not in wb.sheetnames:
                    print(f"  ✗ La hoja '{nombre_hoja}' no existe.")
                    continue

                ws = wb[nombre_hoja]
                
                # Obtener el estilo base de A1 una sola vez fuera de los bucles para mejorar rendimiento
                estilo_base = ws['A1']._style

                # Limpieza masiva e iterativa celda por celda
                for fila in ws[rango_borrar]:
                    for celda in fila:
                        celda.value = None
                        celda._style = estilo_base
                        celda.comment = None
                        celda.hyperlink = None

                print(f"  ✓ Rango {rango_borrar} borrado a fondo en hoja: {nombre_hoja}")

            except Exception as e_hoja:
                print(f"  ⚠ No se pudo limpiar '{nombre_hoja}'. Error: {e_hoja}")

        print("Guardando archivo en el disco de forma segura...")
        fd, temp_file = tempfile.mkstemp(prefix="Base_FONAFE_", suffix=".xlsm", dir=os.path.dirname(ruta_archivo))
        os.close(fd)
        
        wb.save(temp_file)
        wb.close()  # Cierre explícito para liberar la RAM de las 72,000 celdas
        os.replace(temp_file, ruta_archivo)
        print("✓ Archivo guardado correctamente sin bloquear el flujo principal.")

    except PermissionError:
        print("✗ Error: El archivo está abierto. Ciérrelo en Excel de escritorio antes de ejecutar.")
    except Exception as e:
        print(f"✗ Error crítico en el hilo de openpyxl: {e}")
        
    print("Proceso de desinfección y formateo openpyxl finalizado.")
