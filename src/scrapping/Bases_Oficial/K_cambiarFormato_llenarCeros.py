import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def obtener_ultima_fila(worksheet, columna='A', fila_minima=10):
    ultima_fila_encontrada = 1
    for fila in range(worksheet.max_row, 1, -1):
        valor = worksheet[f'{columna}{fila}'].value
        if valor is not None and str(valor).strip() != '':
            ultima_fila_encontrada = fila
            break
    return max(fila_minima, ultima_fila_encontrada)

def agregar_formulas_fbk_presupuesto(worksheet, ultima_fila):
    for fila in range(5, ultima_fila + 1):
        worksheet[f"AE{fila}"] = f"=SUM(OFFSET(F{fila},0,1,1,12))"
        worksheet[f"AF{fila}"] = f"=SUM(OFFSET(F{fila},0,1,1,'Sistema Cierre'!$C$4))"
        worksheet[f"AG{fila}"] = f"=SUM(OFFSET(R{fila},0,1,1,'Sistema Cierre'!$C$4))"

def cambiar_formato_y_llenar_ceros(worksheet):
    ultima_fila = obtener_ultima_fila(worksheet, columna='A', fila_minima=10)
    
    agregar_formulas_fbk_presupuesto(worksheet, ultima_fila)
    
    for fila in range(5, ultima_fila + 1):
        worksheet.cell(row=fila, column=2).alignment = Alignment(horizontal='center')
        
        for col in range(6, 34):
            celda = worksheet.cell(row=fila, column=col)
            celda.number_format = '#,##0'
            
            if 7 <= col <= 30:
                if celda.value is None or str(celda.value).strip() == '':
                    celda.value = 0
                    
    return ultima_fila

def agregar_formulas_fbk_fc(worksheet, ultima_fila):
    for fila in range(5, ultima_fila + 1):
        worksheet[f"AD{fila}"] = f"=SUM(OFFSET(E{fila},0,1,1,12))"
        worksheet[f"AE{fila}"] = f"=SUM(OFFSET(E{fila},0,1,1,'Sistema Cierre'!$C$4))"
        worksheet[f"AF{fila}"] = f"=SUM(OFFSET(Q{fila},0,1,1,'Sistema Cierre'!$C$4))"

def cambiar_formato_y_llenar_ceros_fc(worksheet):
    ultima_fila = obtener_ultima_fila(worksheet, columna='A', fila_minima=10)
    
    agregar_formulas_fbk_fc(worksheet, ultima_fila)
    
    for fila in range(5, ultima_fila + 1):
        worksheet.cell(row=fila, column=2).alignment = Alignment(horizontal='center')
        
        for col in range(6, 33):
            celda = worksheet.cell(row=fila, column=col)
            celda.number_format = '#,##0'
            
            if 6 <= col <= 29:
                if celda.value is None or str(celda.value).strip() == '':
                    celda.value = 0
                    
    return ultima_fila

def cambiar_formato(ruta_principal):
    ruta_excel = rf"{ruta_principal}\Base FONAFE WEB al mes.xlsm"
    
    if not os.path.exists(ruta_excel):
        print(f"Error: No se encontro el archivo en la ruta: {ruta_excel}")
        return

    print("Cargando archivo Excel...")
    
    try:
        import tempfile
        workbook = load_workbook(ruta_excel, keep_vba=True)
        
        hojas_validas = True
        if "FBK-Alineado PRE" not in workbook.sheetnames:
            print("Error: La hoja 'FBK-Alineado PRE' no existe en el archivo.")
            hojas_validas = False
        if "FBK-Alineado FLU" not in workbook.sheetnames:
            print("Error: La hoja 'FBK-Alineado FLU' no existe en el archivo.")
            hojas_validas = False
            
        if not hojas_validas:
            return

        print("Procesando hoja FBK-Alineado PRE...")
        ws_pre = workbook["FBK-Alineado PRE"]
        filas_pre = cambiar_formato_y_llenar_ceros(ws_pre)
        print(f"Completo FBK alineado al Presupuesto. Procesadas {filas_pre} filas.")

        print("Procesando hoja FBK-Alineado FLU...")
        ws_flu = workbook["FBK-Alineado FLU"]
        filas_flu = cambiar_formato_y_llenar_ceros_fc(ws_flu)
        print(f"Completo FBK alineado al FC. Procesadas {filas_flu} filas.")
        
        print("Guardando cambios de forma segura...")
        fd, temp_file = tempfile.mkstemp(prefix="Base_FONAFE_", suffix=".xlsm", dir=os.path.dirname(ruta_excel))
        os.close(fd)
        
        workbook.save(temp_file)
        workbook.close()
        os.replace(temp_file, ruta_excel)
        print("Proceso completo de ambas hojas ejecutado exitosamente.")

    except PermissionError:
        print("Error: No se pudo guardar el archivo. Por favor, asegurese de cerrarlo en Excel.")
    except Exception as e:
        print(f"Ocurrio un error inesperado: {e}")
