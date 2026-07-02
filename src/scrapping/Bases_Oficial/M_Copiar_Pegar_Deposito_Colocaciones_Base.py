import os
import asyncio
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Font

def obtener_ultima_fila(ws, columna='C', fila_minima=2):
    print(f"Buscando última fila... (Excel reporta un total de {ws.max_row} filas totales)")
    for fila in range(ws.max_row, fila_minima - 1, -1):
        valor = ws[f"{columna}{fila}"].value
        if valor is not None and str(valor).strip() != '':
            return fila
    return fila_minima


# Convertida a función principal asíncrona
async def copiar_pegar_deposiciones_colocaciones(ruta_principal):
    ruta_origen = os.path.join(ruta_principal, "EJECUCION", "Depositos_Colocaciones_Ejecucion.xlsx")
    ruta_destino = os.path.join(ruta_principal, "Base FONAFE WEB al mes.xlsm")

    if not os.path.exists(ruta_origen):
        print("No existe archivo origen.")
        return

    if not os.path.exists(ruta_destino):
        print("No existe archivo destino.")
        return

    # Subfunción síncrona interna que correrá dentro del Pool de Hilos
    def _procesar_sync():
        wb_origen = None
        wb_destino = None

        try:
            print("Abriendo archivo origen (esto puede tomar unos segundos)...")
            # data_only=True asegura leer los valores calculados y no las fórmulas de origen
            wb_origen = load_workbook(ruta_origen, data_only=True)
            ws_origen = wb_origen.active

            print("Abriendo archivo destino (esto puede tomar más tiempo por las macros)...")
            wb_destino = load_workbook(ruta_destino, keep_vba=True)
            ws_destino = wb_destino["Depósitos y Colocaciones"]

            ultima_fila = obtener_ultima_fila(ws_origen, columna='C', fila_minima=2)
            print(f"La última fila real con datos es: {ultima_fila}")

            columnas_origen = [
                'C', 'E', 'G', 'M',
                'R', 'T', 'I', 'V',
                'W', 'U', 'Q', 'O'
            ]
            
            fuente_normal = Font(color="000000", italic=False, bold=False)
            fila_destino = 5

            print("Copiando y pegando datos en segundo plano...")
            for fila_origen in range(2, ultima_fila + 1):
                for col_idx, col_letra in enumerate(columnas_origen, start=1):
                    valor = ws_origen[f"{col_letra}{fila_origen}"].value

                    celda_destino = ws_destino.cell(row=fila_destino, column=col_idx)
                    celda_destino.value = valor
                    celda_destino.font = fuente_normal 

                fila_destino += 1

            print("¡Datos pegados! Iniciando proceso de GUARDADO...")
            print(">>> ATENCIÓN: El guardado puede demorar varios minutos. Por favor, espera... <<<")
            
            temp_file = ruta_destino + ".tmp"
            wb_destino.save(temp_file)
            wb_destino.close()
            wb_destino = None
            os.replace(temp_file, ruta_destino)
            print(f"\n✓ ¡Proceso completado con éxito! Filas copiadas: {ultima_fila - 1}")

        except PermissionError:
            print("\n✗ ERROR: El archivo destino está abierto. Ciérrelo en Excel antes de continuar.")
        except Exception as e:
            print(f"\n✗ ERROR Inesperado en la transferencia: {e}")
        finally:
            print("Liberando punteros de archivos de la memoria...")
            if wb_origen:
                wb_origen.close()
            if wb_destino:
                wb_destino.close()

    # --- COORDINACIÓN ASÍNCRONA ---
    loop = asyncio.get_running_loop()
    
    with ThreadPoolExecutor(max_workers=1) as executor:
        await loop.run_in_executor(executor, _procesar_sync)
        
    print("Proceso de transferencia openpyxl finalizado por completo.")


