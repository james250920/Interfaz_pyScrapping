import os
import asyncio
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook

def obtener_ultima_fila(worksheet, columna='C', fila_minima=10):
    ultima_fila_encontrada = 1
    
    for fila in range(worksheet.max_row, 1, -1):
        valor = worksheet[f'{columna}{fila}'].value
        if valor is not None and str(valor).strip() != '':
            ultima_fila_encontrada = fila
            break
            
    return max(fila_minima, ultima_fila_encontrada)

def agregar_formulas_fbk_fc(worksheet):
    ultima_fila = obtener_ultima_fila(worksheet, columna='C', fila_minima=10)

    for fila in range(5, ultima_fila + 1):
        worksheet[f"AD{fila}"] = f"=SUM(OFFSET(E{fila},0,1,1,12))"
        worksheet[f"AE{fila}"] = f"=SUM(OFFSET(E{fila},0,1,1,'Sistema Cierre'!$C$4))"
        worksheet[f"AF{fila}"] = f"=SUM(OFFSET(Q{fila},0,1,1,'Sistema Cierre'!$C$4))"

    return ultima_fila

def agregar_formulas_fbk_presupuesto(worksheet):
    ultima_fila = obtener_ultima_fila(worksheet, columna='C', fila_minima=10)

    for fila in range(5, ultima_fila + 1):
        worksheet[f"AE{fila}"] = f"=SUM(OFFSET(F{fila},0,1,1,12))"
        worksheet[f"AF{fila}"] = f"=SUM(OFFSET(F{fila},0,1,1,'Sistema Cierre'!$C$4))"
        worksheet[f"AG{fila}"] = f"=SUM(OFFSET(R{fila},0,1,1,'Sistema Cierre'!$C$4))"

    return ultima_fila


# Convertida a función principal asíncrona
async def desplegar_formulas(ruta_principal):
    ruta_excel = os.path.join(ruta_principal, "Base FONAFE WEB al mes.xlsm")

    if not os.path.exists(ruta_excel):
        print(f"Error: No se encontró el archivo en la ruta: {ruta_excel}")
        return

    # Subfunción interna síncrona que correrá dentro del Executor
    def _desplegar_sync():
        print("Cargando archivo Excel en segundo plano...")
        try:
            # keep_vba=True es vital para no corromper tu .xlsm
            workbook = load_workbook(ruta_excel, keep_vba=True)
            
            hojas_validas = True
            if "FBK-Alineado FLU" not in workbook.sheetnames:
                print("Error: La hoja 'FBK-Alineado FLU' no existe.")
                hojas_validas = False
                
            if "FBK-Alineado PRE" not in workbook.sheetnames:
                print("Error: La hoja 'FBK-Alineado PRE' no existe.")
                hojas_validas = False
                
            if not hojas_validas:
                workbook.close()
                return

            print("Procesando hoja FBK-Alineado FLU...")
            ws_flu = workbook["FBK-Alineado FLU"]
            filas_flu = agregar_formulas_fbk_fc(ws_flu)
            print(f"  ✓ Fórmulas agregadas en FLU hasta la fila {filas_flu}.")

            print("Procesando hoja FBK-Alineado PRE...")
            ws_pre = workbook["FBK-Alineado PRE"]
            filas_pre = agregar_formulas_fbk_presupuesto(ws_pre)
            print(f"  ✓ Fórmulas agregadas en PRE hasta la fila {filas_pre}.")
            
            print("Guardando cambios en el disco duro...")
            workbook.save(ruta_excel)
            workbook.close()  # Cierre explícito para liberar el flujo de memoria
            print("✓ Proceso completo exitosamente.")

        except PermissionError:
            print("Error: No se pudo guardar. Asegúrate de cerrar el archivo en Excel de escritorio.")
        except Exception as e:
            print(f"Ocurrió un error inesperado en la escritura de fórmulas: {e}")

    # --- COORDINACIÓN ASÍNCRONA ---
    loop = asyncio.get_running_loop()
    
    with ThreadPoolExecutor(max_workers=1) as executor:
        await loop.run_in_executor(executor, _desplegar_sync)
