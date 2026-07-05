import os
import time
import tempfile
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def actualizar_marco(ruta: str):

    carpeta = Path(ruta).expanduser().resolve() / "MARCO"

    CATEGORIA_ORDEN = {"M": 1, "I": 2, "P": 3, "S": 4, "3": 5}

    def validar_temporal_excel(ruta_temporal: Path, keep_vba: bool) -> None:
        """
        Verifica que el archivo temporal se haya guardado correctamente
        antes de reemplazar el archivo original.
        """

        if not ruta_temporal.exists():
            raise FileNotFoundError(f"No se generó el archivo temporal: {ruta_temporal}")

        if ruta_temporal.stat().st_size <= 0:
            raise ValueError(f"El archivo temporal está vacío: {ruta_temporal}")

        wb_validacion = None

        try:
            wb_validacion = load_workbook(
                filename=str(ruta_temporal),
                keep_vba=keep_vba,
                read_only=True,
            )
        finally:
            if wb_validacion is not None:
                wb_validacion.close()

    def procesar_archivo(ruta_archivo: Path) -> dict:
        archivo = ruta_archivo.name

        resultado = {
            "archivo": archivo,
            "ok": False,
            "hoja": "",
            "eliminadas": 0,
            "error": "",
        }

        if not ruta_archivo.exists():
            resultado["error"] = f"El archivo no existe: {ruta_archivo}"
            return resultado

        if not ruta_archivo.is_file():
            resultado["error"] = f"La ruta no es un archivo válido: {ruta_archivo}"
            return resultado

        extension = ruta_archivo.suffix.lower()

        if extension == ".xls":
            resultado["error"] = (
                "Formato .xls no soportado por openpyxl. "
                "Debe convertirse previamente a .xlsx o .xlsm."
            )
            return resultado

        if extension not in {".xlsx", ".xlsm"}:
            resultado["error"] = f"Extensión no soportada: {extension}"
            return resultado

        keep_vba = extension == ".xlsm"

        for intento in range(3):
            wb = None
            temp_file = None

            try:
                df = pd.read_excel(
                    str(ruta_archivo),
                    sheet_name=0,
                    header=0,
                    engine="openpyxl",
                ).copy()

                if len(df.columns) <= 5:
                    raise ValueError(
                        "El archivo no tiene suficientes columnas. "
                        "Se requieren al menos 6 columnas para usar C y F."
                    )

                col_empresa = df.columns[2]       # columna C
                col_modificacion = df.columns[5]  # columna F

                df[col_empresa] = df[col_empresa].astype(str).str.strip()
                df[col_modificacion] = df[col_modificacion].astype(str).str.strip()

                df["_orden"] = df[col_modificacion].map(CATEGORIA_ORDEN).fillna(0)

                idx_max = df.groupby(col_empresa)["_orden"].idxmax()

                mejor_mod = (
                    df.loc[idx_max, [col_empresa, col_modificacion]]
                    .set_index(col_empresa)[col_modificacion]
                    .copy()
                )

                mascara = df.apply(
                    lambda r: mejor_mod.get(r[col_empresa], "") == r[col_modificacion],
                    axis=1,
                )

                eliminadas = int((~mascara).sum())

                df_filtrado = (
                    df[mascara]
                    .drop(columns=["_orden"])
                    .reset_index(drop=True)
                    .copy()
                )

                wb = load_workbook(
                    filename=str(ruta_archivo),
                    keep_vba=keep_vba,
                )

                nombre_copia = f"Formulacion_{len(wb.worksheets) + 1}"
                ws_copia = wb.create_sheet(title=nombre_copia)

                ws_copia.append(list(df_filtrado.columns))

                for row in df_filtrado.itertuples(index=False, name=None):
                    ws_copia.append(list(row))

                fd, temp_name = tempfile.mkstemp(
                    prefix=f".{ruta_archivo.stem}_",
                    suffix=extension,
                    dir=str(ruta_archivo.parent),
                )
                os.close(fd)

                temp_file = Path(temp_name)

                wb.save(str(temp_file))

                # Cierre explícito del workbook antes de reemplazar el archivo original.
                wb.close()
                wb = None

                validar_temporal_excel(temp_file, keep_vba=keep_vba)

                os.replace(str(temp_file), str(ruta_archivo))
                temp_file = None

                resultado.update(
                    {
                        "ok": True,
                        "hoja": nombre_copia,
                        "eliminadas": eliminadas,
                        "error": "",
                    }
                )

                return resultado

            except PermissionError:
                resultado["error"] = "Archivo bloqueado en disco"
                time.sleep(0.5)

            except Exception as e:
                resultado["error"] = str(e)
                return resultado

            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception as e:
                        logger.warning(
                            "No se pudo cerrar correctamente el workbook %s: %s",
                            ruta_archivo,
                            e,
                        )

                if temp_file is not None:
                    try:
                        if temp_file.exists():
                            temp_file.unlink()
                    except Exception as e:
                        logger.warning(
                            "No se pudo eliminar el archivo temporal %s: %s",
                            temp_file,
                            e,
                        )

        return resultado

    if not carpeta.exists():
        logger.error("La carpeta no existe: %s", carpeta)
        return

    if not carpeta.is_dir():
        logger.error("La ruta no es una carpeta válida: %s", carpeta)
        return

    archivos_unicos = sorted(
        {
            archivo.resolve()
            for archivo in carpeta.iterdir()
            if archivo.is_file()
            and archivo.suffix.lower() in {".xlsx", ".xls", ".xlsm"}
            and not archivo.name.startswith("Gastos_Capital")
            and not archivo.name.startswith("~$")
        }
    )

    if not archivos_unicos:
        logger.info("No se encontraron archivos Excel válidos para procesar.")
        return

    total = len(archivos_unicos)
    completados = 0
    errores = 0

    t_inicio = time.perf_counter()

    logger.info("Archivos únicos a procesar secuencialmente: %s", total)

    for ruta_archivo in archivos_unicos:
        res = procesar_archivo(ruta_archivo)

        completados += 1
        estado = f"[{completados}/{total}]"

        if res["ok"]:
            logger.info(
                "  %s OK  %s  ->  '%s' | eliminadas: %s",
                estado,
                res["archivo"],
                res["hoja"],
                res["eliminadas"],
            )
        else:
            errores += 1
            logger.error(
                "  %s ERR %s  ->  %s",
                estado,
                res["archivo"],
                res["error"],
            )
            
    t_fin = time.perf_counter()

    logger.info(
        "Completado en %.2fs | OK: %s | Errores: %s",
        t_fin - t_inicio,
        completados - errores,
        errores,
    )