import os
import time
import openpyxl
from openpyxl.utils import range_boundaries
import tempfile
import shutil

def copiar_pegar(ruta_principal, anio, mes):
    inicio = time.time()

    RUTA_DESTINO = os.path.join(ruta_principal, "Eva 2026 - Información al mes.xlsm")
    RUTA_ORIGEN = os.path.join(ruta_principal, "Base FONAFE WEB al mes.xlsm")

    OPERACIONES = [
        ("PRE ", "C4:Z8000", "BPRE", "AG4:BD8000"),
        ("FLU", "C4:Z8000", "BFLU", "AG4:BD8000"),
        ("ESF", "C4:Z8000", "BESF", "AG4:BD8000"),
        ("ERI", "C4:Z8000", "BERI", "AG4:BD8000"),
    ]

    wb_origen = None
    wb_destino = None

    try:
        if not os.path.exists(RUTA_ORIGEN):
            raise FileNotFoundError(RUTA_ORIGEN)

        if not os.path.exists(RUTA_DESTINO):
            raise FileNotFoundError(RUTA_DESTINO)

        print("Cargando archivo origen en memoria (esto puede tardar)...")
        # data_only=True extrae los VALORES y no las fórmulas
        wb_origen = openpyxl.load_workbook(RUTA_ORIGEN, data_only=True, read_only=True)

        print("Cargando archivo destino en memoria...")
        # keep_vba=True es CRÍTICO para no destruir el .xlsm
        wb_destino = openpyxl.load_workbook(RUTA_DESTINO, keep_vba=True)

        # 1. Escritura de Cabecera
        ws_bpre = wb_destino["BPRE"]
        ws_bpre["A2"] = anio
        ws_bpre["B2"] = mes

        # 2. Transferencia masiva
        for hoja_o, rng_o, hoja_d, rng_d in OPERACIONES:
            ws_o = wb_origen[hoja_o]
            ws_d = wb_destino[hoja_d]

            # Calculamos las coordenadas numéricas de los rangos (Ej: A1 -> col=1, row=1)
            min_col_o, min_row_o, max_col_o, max_row_o = range_boundaries(rng_o)
            min_col_d, min_row_d, max_col_d, max_row_d = range_boundaries(rng_d)

            print(f"Transfiriendo {hoja_o} → {hoja_d}...")
            
            # iter_rows con values_only=True es la forma más rápida de leer en openpyxl
            datos_origen = ws_o.iter_rows(min_row=min_row_o, max_row=max_row_o,
                                          min_col=min_col_o, max_col=max_col_o, 
                                          values_only=True)

            # Insertamos fila por fila y celda por celda en el destino
            for i_row, fila in enumerate(datos_origen):
                for i_col, valor in enumerate(fila):
                    # Solo escribimos si el valor no es nulo para ganar algo de velocidad
                    if valor is not None:
                        ws_d.cell(row=min_row_d + i_row, column=min_col_d + i_col, value=valor)

            print(f"  ✓ {hoja_o} completado.")

        # 3. Guardado Atómico
        print("Guardando archivo temporal...")
        fd, ruta_temp = tempfile.mkstemp(prefix="Eva_2026_", suffix=".xlsm", dir=os.path.dirname(RUTA_DESTINO))
        os.close(fd)
        
        wb_destino.save(ruta_temp)
        wb_destino.close()
        wb_destino = None

        os.replace(ruta_temp, RUTA_DESTINO)
        print("✓ Archivo maestro guardado y reemplazado correctamente.")

    except Exception as e:
        print(f"\n✗ ERROR EN PROCESO OPENPYXL:\n{e}")

    finally:
        # En openpyxl siempre debemos invocar el cierre, especialmente con read_only=True
        if wb_origen:
            try:
                wb_origen.close()
            except:
                pass
        if wb_destino:
            try:
                wb_destino.close()
            except:
                pass

    print(f"Tiempo total: {round(time.time() - inicio, 2)}s")